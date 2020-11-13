# -*- coding: utf-8

import json
import re
import difflib
import openpyxl
import os
import Levenshtein
from collections import Counter
from wasabi import color


class jj2_assert(object):
    """

    """

    def __init__(self, members="401-2.txt"):
        """
        共通変数
        """
        self.scores = {}
        self.trans = str.maketrans({
            "：": ":",
            "、": ",",
            "，": ",",
            "。": ".",
            "．": ".",
            "　": "",
            "”": "\"",
            "＞": ">",
            "＜": "<",
            "０": "0",
            "１": "1",
            "＼": "\\",
            "ー": "-",
            "（": "(",
            "）": ")",
            "＝": "=",
            "！": "!",
            "〜": "~",
            "＠": "@",
            "＃": "#",
            "＄": "$",
            "％": "%",
            "＆": "&",
            "＾": "^",
            "＊": "*",
            "＋": "+",
            "｜": "|",
            "｛": "{",
            "｝": "}",
            "；": ";",
            "？": "?",
            "／": "/",
        })
        self.pattern = "(?=\d{2}10370\d{3})"
        self.comments = {}
        ext = os.path.splitext(members)[1]
        if ext == ".xlsx":
            self.load_members_xl(members)
        else:
            self.load_members(members)
        self.skip_status = ["未", "RE", "CE"]
        self.answer_master = {}

    def translate(self, input1):
        """
        表示向けにややこしい文字を変換する
        """
        return input1.strip().translate(self.trans).replace(" ", "")

    def reformat(self, input_str):
        """
        正誤判定向けにややこしい文字を変換する
        """
        input_str = input_str.strip().translate(self.trans)
        spliter = [
            "\n",
            "\r\n",
            " ",
            "　",
            "\t"
        ]
        p = "|".join(spliter)
        return re.sub(p, "", input_str)

    def load_members(self, filepath):
        """
        メンバーを読み込む
        """
        with open(filepath) as f:
            strings = f.read().split("\n")
        for s in strings:
            if not s:
                continue
            self.scores[s] = {"id": s, "pre": "", "score": ""}

    def load_scores(self, filepath):
        """
        途中採点結果を読み込む
        """
        with open(filepath) as f:
            self.scores = json.load(f)

    def write_scores(self, filepath="output.json"):
        """
        採点結果を書き出す
        """
        with open(filepath, "w") as f:
            json.dump(self.scores, f, indent=2, ensure_ascii=False)

    def no_submitted(self, filepath, filename, delay=False):
        """
        未提出者を判定する
        """
        # print("未提出者読み込み")
        with open(filepath) as f:
            strings = f.read()
        files = re.split("(?=\n\S+\.java 未提出者)",
                         strings)
        m_file = {}
        for f in files:
            num = f.strip().find(".java")
            if num > 1:
                m_file[f.strip()[:num] + ".java"] = f.strip()[num + 5:]

        if filename not in m_file:
            return
        # print(m_file[filename])
        no_submitted = re.split(self.pattern, m_file[filename])
        for n in no_submitted:
            if n[:10] in self.scores.keys():
                self.score(n[:10], "未", delay)

    def runtime_error(self, filepath, delay=False):
        """
        実行エラーを判定する
        """
        # print("エラーログ読み込み")
        with open(filepath) as f:
            strings = f.read()
        r_error = re.split(self.pattern, strings)
        for n in r_error:
            if n[:10] in self.scores.keys():
                self.score(n[:10], "RE", delay)

    def comment_check(self, filepath, filename):
        """
        コメントチェックをする(未実装)
        """
        with open(filepath) as f:
            strings = f.read()
        files = re.split("(?=-----\S+\.java)-----", strings)
        m_file = {}
        for f in files:
            num = f.strip().find(".java")
            if num >= 1:
                m_file[f.strip()[:num] + ".java"] = f.strip()[num + 5:]
        if filename not in m_file:
            return
        comments = re.split(self.pattern, m_file[filename])
        for n in comments:
            if len(n) < 10:
                continue
            if n[:10] in self.scores.keys():
                self.comments[n[:10]] = n.split("...")[1].strip()

    def diff(self, i1, i2):
        """
        差分をとる
        """
        output = []
        a = self.translate(i1)
        b = self.translate(i2)
        matcher = difflib.SequenceMatcher(
            None, a, b)
        # print(i1)
        # print(i2)
        for opcode, a0, a1, b0, b1 in matcher.get_opcodes():
            if opcode == "equal":
                output.append(a[a0:a1])
            elif opcode == "insert":
                output.append(color(b[b0:b1], fg=16, bg="green"))
            elif opcode == "delete":
                output.append(color(a[a0:a1], fg=16, bg="red"))
            elif opcode == "replace":
                output.append(color(b[b0:b1], fg=16, bg="green"))
                output.append(color(a[a0:a1], fg=16, bg="red"))
        return "".join(output)

    def ask(self):
        """
        y/n入力を受け付ける
        """
        while True:
            score = input("[OK(o)/WA(n)] >> ")
            if score == "o":
                score = "OK"
            elif score == "n":
                score = "WA"
            if score in ["OK", "WA"]:
                break
        return score

    def score(self, number, score, delay):
        """
        採点を受け付けるインターフェース
        """
        target = self.scores[number]
        if not delay:
            if target["pre"] == "":
                target["pre"] = score
            elif target["pre"] == "OK" and score != "OK":
                target["pre"] = score
            else:
                #print("pre  {}({}:{})".format(number, target["pre"], score))
                pass
        else:
            if target["pre"] == "OK":
                target["score"] = "OK"
            elif target["score"] == "":
                target["score"] = score
            elif target["score"] == "OK" and score != "OK":
                target["score"] = score
            else:
                #print("post {}({}:{})".format(number, target["score"], score))
                pass

    def print_status(self):
        """
        現在の状況を表示する
        """
        for k, v in self.scores.items():
            print("{},{},{}".format(k, v["pre"], v["score"]))

    def write_excel(self, kadai, book="401-2.xlsx", output="output.xlsx", resubmit="resubmit.xlsx"):
        """
        Excelシートに書き込む
        """
        wb = openpyxl.load_workbook(book)
        ws = wb["MT_kadai"]
        rows = {}
        for row in ws["C4:C275"]:
            if row[0].value:
                rows[str(row[0].value)] = row[0].row

        ws["F3"].value = kadai

        resubmit_wb = openpyxl.load_workbook(resubmit)
        resubmit_ws = resubmit_wb["Sheet1"]
        resubmit_counter = 2
        for k, v in self.scores.items():
            if k in rows.keys():
                cell = ws.cell(rows[k], 6)
                if v["score"] == "OK" and v["pre"] != "OK":
                    cell.value = "遅"
                elif v["score"] == "OK" and v["pre"] == "OK":
                    continue
                elif v["score"] == "" or v["pre"] == "":
                    print("Score Not Found({}, {})", k, v)
                else:
                    cell.value = v["score"]
                    room, name = self._get_row_info(ws, rows[k])
                    self._write_num(resubmit_ws, resubmit_counter,
                                    room, k, name, kadai, v["score"])
                    resubmit_counter += 1

        wb.save(output)
        resubmit_wb.save("re-" + output)

    def load_members_xl(self, book="401-2.xlsx"):
        """
        Excelシートから読み込む
        """
        wb = openpyxl.load_workbook(book)
        ws = wb["MT_kadai"]
        for row in ws["C4:C275"]:
            if row[0].value:
                s = str(row[0].value)
                self.scores[s] = {"id": s, "pre": "", "score": ""}

    def _get_row_info(self, ws, row):
        """
        Excelシートの指定位置から情報を受け取る
        """
        r = ws["A4:F275"]
        c = r[row-4]
        # room_number, name
        return c[0].value, c[3].value

    def _write_num(self, ws, row, room, number, name, kadai, score):
        """
        Excelシートに情報を書き込む
        """
        ws.cell(row, 1).value = room
        ws.cell(row, 2).value = number
        ws.cell(row, 3).value = name
        ws.cell(row, 4).value = kadai
        ws.cell(row, 5).value = score

    def run_v2(self, pattern="pre", kadai="5-3", numbers=[1, 2, 3], classname="MainForMetabolickChecker", root="05", rooms=[401, 402]):
        """
        pre/postの採点を一巡行う
        """
        delay = (pattern == "post")
        print(root, pattern, classname)
        # 未提出確認
        self.no_submitted(os.path.join(
            root, pattern + "/miteishutu.txt"), classname+".java", delay=delay)
        for n in [str(i) for i in numbers]:
            print('''
        # ================
        # {} {}:{}
        # ================
        '''.format(pattern, kadai, n))
            for room in [str(i) for i in rooms]:
                # Runtime Error
                self.runtime_error(os.path.join(
                    root, pattern, room, classname + n + ".err"), delay=delay)
            # 採点
            self.scoring_v2(os.path.join(root, pattern, str(rooms[0]), classname+n+".log"),
                            os.path.join(root, pattern, str(rooms[1]), classname+n+".log"), os.path.join(root, "ans{}_{}.txt".format(kadai, n)), delay=delay)

    def scoring_v2(self, NOTE_FILE, NOTE_FILE2, ANSWER_FILE, delay):
        """
        Answerと比較して正誤を決め、記録する
        """
        logs = {}
        with open(ANSWER_FILE) as f:
            answer = f.read()
        with open(NOTE_FILE) as f:
            note_str = f.read()
        for s in re.split(self.pattern, note_str):
            if not s:
                continue
            logs[s[:10]] = s[10:].strip()
        with open(NOTE_FILE2) as f:
            note_str = f.read()
        for s in re.split(self.pattern, note_str):
            if not s:
                continue
            logs[s[:10]] = s[10:].strip()

        if delay:
            keys = [self.translate(logs[k].strip()) for k in self.scores.keys() if self.scores[k]
                    ["score"] not in self.skip_status and self.scores[k]["pre"] != "OK"]
            # student_id_list = [k for k in self.scores.keys() if self.scores[k]
            #                   ["score"] not in self.skip_status and self.scores[k]["pre"] != "OK"]
        else:
            keys = [self.translate(logs[k].strip()) for k in self.scores.keys()
                    if self.scores[k]["pre"] not in self.skip_status]
            # student_id_list = [k for k in self.scores.keys()
            #                   if self.scores[k]["pre"] not in self.skip_status]
        # print(keys)
        counter = dict(Counter(keys))
        counter.update(self.answer_master)
        f_answer = self.reformat(answer)

        for k, v in zip(counter.items()):
            if type(v) == str:  # OK or NG
                continue
            if f_answer in self.reformat(k):
                counter[k] = "OK"
            else:
                print("----- {}件 -----".format(counter[k]))
                print(self.diff(answer, k).strip())
                counter[k] = self.ask()
        # print(counter)
        for k, v in logs.items():
            if self.scores[k]["pre"] == "OK" and delay:
                self.score(k, "OK", delay)
                self.print_score(k, delay)
                continue
            if self.translate(v.strip()) not in counter:
                # print(v)
                self.print_score(k, delay)
                continue
            if counter[self.translate(v.strip())]:
                self.score(k, counter[self.translate(v.strip())], delay)
                self.print_score(k, delay)
            else:
                self.print_score(k, delay)
                print("no_update")
        self.answer_master = counter

    def print_score(self, student_id, delay):
        """
        一人分の情報を表示する
        """
        if delay:
            print("{}: {}".format(student_id,
                                  self.scores[student_id]["score"]))
        else:
            print("{}: {}".format(student_id,
                                  self.scores[student_id]["pre"]))


def levenstein(str1, str2):
    """
    Levenstein距離(編集距離)を算出
    """
    lev_dist = Levenshtein.distance(str1, str2)
    divider = len(str1) if len(str1) > len(str2) else len(str2)
    lev_dist = lev_dist / divider
    lev_dist = 1 - lev_dist
    return lev_dist


def jaro_winkler(str1, str2):
    """
    JaroWinkler距離を算出
    """
    jaro_dist = Levenshtein.jaro_winkler(str1, str2)
    return jaro_dist


if __name__ == "__main__":

    members = "401-2.txt"
    kadai = "5-2"
    numbers = [1]
    classname = "GraduationCheck"
    root = "05"
    rooms = [401, 402]
    book = "401-2.xlsx"
    output = "output5-2.xlsx"

    # load_instance
    obj = jj2_assert(members)
    # 途中から始める場合
    # obj.load_scores("5-1pre.json")

    # pre
    obj.run("pre", kadai=kadai, numbers=numbers,
            classname=classname, root=root, rooms=rooms)

    # 一時保存
    # obj.print_status()
    # obj.write_scores("{}pre.json".format(kadai))

    # post
    obj.run("post", kadai=kadai, numbers=numbers,
            classname=classname, root=root, rooms=rooms)

    # ================
    # 終了処理
    # ================
    print('''
    # ================ #
    # 採点終了          #
    # ================ #
    ''')
    obj.print_status()
    obj.write_scores("{}post.json".format(kadai))
    obj.write_excel(kadai, book, output)

    exit()
